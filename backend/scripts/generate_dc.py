#!/usr/bin/env python3
"""
DC Generation Script for Logogear Portal
Based on the working Python script provided by user
Preserves all images, logos, and formatting from the template
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
import json
from openpyxl.cell.cell import MergedCell
import shutil

def first_free_cell(ws, row_idx, start_col_idx):
    """Return first non-merged cell in given row starting from start_col_idx."""
    col_idx = start_col_idx
    while True:
        cell = ws.cell(row=row_idx, column=col_idx)
        if not isinstance(cell, MergedCell):
            return cell
        col_idx += 1

def generate_dc_files(csv_file_path, template_path, output_dir):
    """
    Generate DC files from CSV data using the template
    Preserves ALL images, logos, and formatting from the original template
    Returns: dict with success status and file paths
    """
    try:
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Verify template exists
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Load CSV
        df = pd.read_csv(csv_file_path)
        print(f"[SUCCESS] Loaded {len(df)} records")
        print("[INFO] Columns:", df.columns.tolist())
        
        current_date = datetime.now()
        date_str = current_date.strftime('%d/%m/%Y')
        date_for_filename = current_date.strftime('%d%m%Y')
        
        generated_files = []
        
        for idx, row in df.iterrows():
            try:
                # Load template for each file - this preserves ALL content including images
                print(f"[PROCESSING] Row {idx + 1}: Loading template...")
                wb = load_workbook(template_path)
                ws = wb.active
                
                # Extract customer data
                full_name = str(row.get('Full Name', ''))
                street = str(row.get('Street Address', ''))
                landmark = str(row.get('Landmark', ''))
                city = str(row.get('City', ''))
                state = str(row.get('State/Province', ''))
                pincode = str(row.get('Postal Code', ''))
                country = str(row.get('Country', 'India'))
                mobile = str(row.get('Mobile Number', ''))
                email = str(row.get('Email', ''))
                
                # Skip if essential data is missing
                if not full_name or not city:
                    print(f"[WARNING] Skipping row {idx + 1}: Missing essential data (Name: '{full_name}', City: '{city}')")
                    continue
                
                print(f"[CUSTOMER] Processing: {full_name} from {city}")
                
                # ===== 1) TO SECTION IN COLUMN C ===== (Exactly like working Python script)
                # Start from column 3 (C), skip merged cells automatically
                print("[DATA] Filling customer data in TO section (Column C)...")
                
                name_cell = first_free_cell(ws, 3, 3)       # row 3, col >= C
                addr1_cell = first_free_cell(ws, 4, 3)      # row 4, col >= C
                landmark_cell = first_free_cell(ws, 5, 3)   # row 5, col >= C
                citypin_cell = first_free_cell(ws, 6, 3)    # row 6, col >= C
                statecty_cell = first_free_cell(ws, 7, 3)   # row 7, col >= C
                mobile_cell = first_free_cell(ws, 8, 3)     # row 8, col >= C
                email_cell = first_free_cell(ws, 9, 3)      # row 9, col >= C
                
                # Fill the data exactly like working Python script
                name_cell.value = full_name
                addr1_cell.value = street
                landmark_cell.value = landmark or ''
                citypin_cell.value = f"{city} - {pincode}"
                statecty_cell.value = f"{state}, {country}"
                mobile_cell.value = f"{mobile}"
                #email_cell.value = f"Email: {email}"
                
                print(f"[SUCCESS] Filled customer data: {full_name}")
                
                # ===== 2) DC NO & DATE ===== (Like working Python script)
                print("[DATE] Setting date...")
                ws['B12'] = date_str   # cell next to "DC No" or "Date"
                
                # ===== 3) OUTPUT FILENAME ===== (Like working Python script)
                safe_name = (str(full_name).replace(' ', '_')
                           .replace('/', '_')
                           .replace('\\', '_')
                           .replace('?', '_')
                           .replace('*', '_')
                           .replace(':', '_')
                           .replace('|', '_')
                           .replace('<', '_')
                           .replace('>', '_')
                           .replace('"', '_'))[:25]
                
                output_filename = f"DC_{safe_name}_{date_for_filename}.xlsx"
                output_file = os.path.join(output_dir, output_filename)
                
                # Save the file - this preserves ALL template content:
                # - Header "Delivery Challan" title
                # - Logogear logo (blue background with white text)
                # - Complete "LOGO GEAR SOLUTION LLP" company details
                # - All table formatting and borders
                # - Circular signature stamp at bottom
                # - All images, logos, and styling
                print(f"[SAVE] Saving file: {output_filename}")
                wb.save(output_file)
                generated_files.append(output_file)
                
                print(f"[SUCCESS] Generated: {output_filename}")
                
            except Exception as e:
                print(f"[ERROR] Error processing row {idx + 1} ({full_name}): {str(e)}")
                continue
        
        print(f"[COMPLETE] Successfully generated {len(generated_files)} DC files!")
        print("[FILES] Generated files:")
        for file_path in generated_files:
            print(f"   - {os.path.basename(file_path)}")
        
        return {
            'success': True,
            'message': f'Generated {len(generated_files)} DC files successfully',
            'files': generated_files,
            'count': len(generated_files)
        }
        
    except Exception as e:
        error_msg = f"Error generating DC files: {str(e)}"
        print(f"[ERROR] {error_msg}")
        return {
            'success': False,
            'message': error_msg,
            'files': [],
            'count': 0
        }

def main():
    """Main function to handle command line arguments"""
    if len(sys.argv) != 4:
        print("Usage: python generate_dc.py <csv_file> <template_file> <output_dir>")
        print("Example: python generate_dc.py data.csv DC-FORMAT.xlsx output/")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    template_file = sys.argv[2]
    output_dir = sys.argv[3]
    
    print("[START] Starting DC Generation...")
    print(f"[CSV] CSV file: {csv_file}")
    print(f"[TEMPLATE] Template: {template_file}")
    print(f"[OUTPUT] Output directory: {output_dir}")
    
    # Validate input files exist
    if not os.path.exists(csv_file):
        print(f"[ERROR] CSV file not found: {csv_file}")
        sys.exit(1)
    
    if not os.path.exists(template_file):
        print(f"[ERROR] Template file not found: {template_file}")
        sys.exit(1)
    
    # Generate DC files
    result = generate_dc_files(csv_file, template_file, output_dir)
    
    # Output result as JSON for Node.js to parse
    print("RESULT_JSON:", json.dumps(result))
    
    if result['success']:
        print("[COMPLETE] DC generation completed successfully!")
        sys.exit(0)
    else:
        print("[FAILED] DC generation failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()